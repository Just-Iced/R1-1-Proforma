import openpyxl.workbook
import json, openpyxl, time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import styles
from area import area
from turfpy.measurement import boolean_point_in_polygon
from geojson import Point, Polygon, Feature
from geopy.geocoders import Nominatim
from selenium.common.exceptions import TimeoutException
# The class to find all heritage properties listed on popular realty websites
class R1_Finder:
    def __init__(self,
                    # Where to store the data files
                    path: str = "data") -> None:
        
        self.path = path
        self.dictionary = {}
        self.parcels = {}
        try:
            self.dictionary = json.load(open(f"{path}/r1_1_properties.json"))
        except FileNotFoundError:
            pass
        
        self.parcels = json.load(open("parcels.json"))
        
        
        firefox_profile = webdriver.FirefoxProfile()
        firefox_profile.set_preference("javascript.enabled", True)
        firefox_profile.set_preference("cookies.enabled", True)
        self.options = webdriver.FirefoxOptions()
        self.options.profile = firefox_profile
        self.geolocator = Nominatim(user_agent="GeoLocator")
        self.r1_zones = []
        self.other_zones = []
        with open("zoning-districts-and-labels.json") as f:
            data = json.load(f)
            for dictionary in data:
                polygon_list = [[]]
                for point in dictionary["geom"]["geometry"]["coordinates"][0]:
                    point.reverse()
                    polygon_list[0].append(tuple(point))
                polygon = Polygon(polygon_list, precision=14)
                if dictionary["zoning_district"] == "R1-1":
                    self.r1_zones.append(polygon)
                else:
                    self.other_zones.append(polygon)
        self.zones = {}
        for area in json.load(open("local-area-boundary.json")):
            polygon_list = [[]]
            for point in area["geom"]["geometry"]["coordinates"][0]:
                point.reverse()
                polygon_list[0].append(tuple(point))
            self.zones[area["name"]] = Polygon(polygon_list, precision=14)
            
        #self.vancouver_api_key = "f6aba1995ad5dcbd2f37c943d36001f1fcfa7441c6243868fa1a0792"
        #self.rentcast_key = "533d316d4fc54e3aac177815ae86d2f6"
        #MultiPolygon([([(-81, 41), (-81, 47), (-72, 47),(-72, 41), (-81, 41)],),([(3.78, 9.28), (-130.91, 1.52), (35.12, 72.234), (3.78, 9.28)],)])
            
    def get_sq_ft(self, address: str) -> float:
        address_data = {}
        try:
            address_data = self.parcels[address.strip()]
        except KeyError:
            return 0
        polygon = address_data["geom"]["geometry"]
        return round(area(polygon) * 10.764, 2)

    def get_lat_lon(self, address: str) -> tuple[float, float]:
        coords = []
        try:
            raw_coords = self.parcels[address]["geo_point_2d"]
            coords = (float(raw_coords["lat"]), float(raw_coords["lon"]))
        except KeyError:
            location = self.geolocator.geocode(f"{address.strip()}, Vancouver", country_codes="CA", timeout=300, namedetails=True)
            if location == None:
                return None
            coords = (float(location.raw['lat']), float(location.raw['lon']))
        return coords

    def is_r1_property(self, address: str) -> bool:
        coords = self.get_lat_lon(address)
        point = Feature(geometry=Point(coords, precision=14))
        for r1_zone in self.r1_zones:
            try:
                if boolean_point_in_polygon(point, r1_zone):
                    for other_zone in self.other_zones:
                        if boolean_point_in_polygon(point, other_zone):
                            return False
                    print(address)
                    return True
            except IndexError:
                return False
        return False
    
    def get_zone(self, address: str) -> str:
        coords = self.get_lat_lon(address)
        point = Feature(geometry=Point(coords, precision=14))
        for zone in self.zones:
            polygon = self.zones[zone]
            if boolean_point_in_polygon(point, polygon):
                return zone
        return ""
    
    def update_realty_data(self) -> dict:
        driver = webdriver.Firefox(options=self.options)
        def _goto_page(pg_num: int) -> None:
            url = f"https://www.realtor.ca/map#ZoomLevel=15&Center=49.2827%2C-123.1207&LatitudeMax=49.25428&LongitudeMax=-123.14113&LatitudeMin=49.19609&LongitudeMin=-123.22857&view=list&CurrentPage={pg_num}&Sort=6-D&PropertyTypeGroupID=1&TransactionTypeId=2&PropertySearchTypeId=0&Currency=CAD&HiddenListingIds=&IncludeHiddenListings=false"
            driver.get(url)
            driver.refresh()
            WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, "listingCardTopBody")))
            
        def _read_page() -> list[tuple]:
            addresses = []
            elements = driver.find_elements(By.CLASS_NAME, "listingCardTopBody")
            for element in elements:
                priceElement = element.find_element(By.CLASS_NAME, "listingCardPrice")
                price = priceElement.get_attribute("title")
                addressElement = element.find_element(By.CLASS_NAME, "listingCardAddress")
                address = addressElement.text
                address = address.split(",")[0].title()
                address_split = address.split()
                numbers = 0
                for word in address_split:
                    try:
                        int(word)
                        numbers += 1
                    except ValueError:
                        break
                if numbers >= 2:
                    address = ""
                    address_split = address_split[1:]
                    for word in address_split:
                        address += f"{word} "
                try:
                    addresses.append((address.strip(), int(price.replace("$","").replace(",",""))))
                except ValueError:
                    pass
            return addresses
        realty_addresses_dict = {}
        for i in range(1, 51):
            print(f"Page {i}")
            try:
                _goto_page(i)
            except TimeoutException:
                break
            time.sleep(2)
            for item in _read_page():
                realty_addresses_dict[item[0]] = {"Price": item[1], "Sqft": self.get_sq_ft(item[0])}
            time.sleep(2)
        driver.close()
        keys = realty_addresses_dict.keys()
        listings = ""
        for key in keys:
            listings += f"{key.strip()}\n"
        with open("data/realty_listings.txt", "w", encoding="utf-8") as f:
            f.write(listings)
        
        return realty_addresses_dict
    
    def get_r1_listings(self) -> None:
        realty_addresses_dict = self.update_realty_data()
        self.dictionary = {}
        for address in realty_addresses_dict.keys():
            if self.is_r1_property(address.rstrip(" \n")):
                self.dictionary[address] = realty_addresses_dict[address]
        
        new_dict = {}
        for key, value in self.dictionary.items():
            if key not in new_dict.keys():
                new_dict[key.removesuffix(" \n").strip()] = value
        self.dictionary = new_dict
        with open(f"{self.path}/r1_1_properties.json", "w") as f:
            json.dump(self.dictionary, f, indent=4)
        
    def generate_proforma(self):
        from time import gmtime, strftime
        from openpyxl.worksheet.hyperlink import Hyperlink
        import os
        cur_day = strftime("%Y-%m-%d", gmtime())
        wb_title = f"Proforma {cur_day}.xlsx"
        wb = openpyxl.load_workbook("Proforma Template.xlsx")
        template = wb["Proforma Template"]
        master_list = wb["Master List"]
        from openpyxl.styles import Font
        format = Font(name="Arial", size=10)
        
        master_row = 5
        for key in self.dictionary:
            sheet = wb.copy_worksheet(template)
            data = self.dictionary[key]
            sheet.title = key
            sheet.cell(1,2).value = key
            sheet.cell(2,2).value = self.get_zone(key)
            sheet.cell(7,4).value = data["Sqft"]
            price = data['Price']
            sheet.cell(45,6).value = price
            
            
            address_cell = master_list.cell(master_row, 1)
            address_cell.value = key
            address_cell.hyperlink = Hyperlink("A1", f'{key}!A1', display=key)
            price_cell = master_list.cell(master_row, 2)
            price_cell.number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
            price_cell.value = price
            
            
            master_row += 1
        wb.save("data_only.xlsx")
        wb.save(wb_title)
        wb.close()
        import xlwings 
        excel_app = xlwings.App(visible=False)
        excel_book = excel_app.books.open("data_only.xlsx")
        excel_book.save()
        excel_app.kill()
        
        wb = openpyxl.load_workbook("data_only.xlsx", data_only=True)
        for key in self.dictionary:
            sheet = wb[key]
            self.dictionary[key]["Profit"] = float(sheet.cell(94,7).internal_value)
        wb.close()
        os.remove("data_only.xlsx") 
        wb = openpyxl.load_workbook(wb_title)
        master_row = 5
        master_list = wb["Master List"]
        for key in self.dictionary:
            profit_cell = master_list.cell(master_row, 3)
            profit_cell.number_format = "[$$-1009]#,##0.00;-[$$-1009]#,##0.00"
            profit = self.dictionary[key]["Profit"]
            profit_cell.value = profit
            if float(profit_cell.value) > 0:
                colour = styles.colors.Color(rgb='11FF00')
            else:
                colour = styles.colors.Color(rgb='00FF0000')
            fill = styles.fills.PatternFill(patternType='solid', fgColor=colour)
            profit_cell.fill = fill
            master_row += 1
        master_list.cell(4,2).font = format
        master_list.cell(4,3).font = format
        try:
            wb.remove(wb["Proforma Template"])
        except KeyError:
            pass
        wb.save(wb_title)

    #Convert the Vancouver parcel JSON into something more usable            
    def convert_parcel_json(self):
        try:
            with open("property-parcel-polygons.json") as f:
                new_dict = {}
                street_conversions = {
                    "St": "Street",
                    "Av": "Avenue",
                }
                
                data = json.load(f)
                for dictionary in data:
                    street_name = str(dictionary['streetname']).title()
                    splt = street_name.split()
                    if splt[-1] in street_conversions:
                        splt[-1] = street_conversions[splt[-1]]
                        street_name = ""
                        for word in splt:
                            street_name += f"{word} "
                    
                    new_dict[f"{dictionary['civic_number']} {street_name.strip()}"] = dictionary
                json.dump(new_dict, open("data/parcels.json", 'w'))
        except FileNotFoundError:
            return
        
        